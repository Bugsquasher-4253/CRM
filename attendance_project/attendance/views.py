import datetime
import os

from django.contrib import messages
from django.contrib.auth import authenticate, login, logout, update_session_auth_hash
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import User
from django.db import IntegrityError, transaction
from django.db.models import Count, Q, Sum
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone

from . import emails as email_service
from .forms import (
    AdminCorrectionResponseForm,
    AdminEmployeeEditForm,
    AdminLeaveResponseForm,
    AdminPasswordChangeForm,
    AdminReimbursementActionForm,
    AdminTicketResponseForm,
    AttendanceCorrectionForm,
    AttendanceRecordForm,
    DepartmentForm,
    EmployeeForm,
    EmployeeProfileEditForm,
    LeaveRequestForm,
    LoginForm,
    ReimbursementForm,
    SalaryForm,
    SalaryStructureForm,
    SupportTicketForm,
    UserForm,
)
from .models import (
    AttendanceBreak,
    AttendanceCorrectionRequest,
    AttendanceRecord,
    Department,
    Employee,
    EmployeeSalaryStructure,
    LeaveRequest,
    Reimbursement,
    ReimbursementAttachment,
    SalaryRecord,
    SupportTicket,
)


def is_admin(user):
    return user.is_staff or user.is_superuser


def _resolve_incomplete_checkins(employee=None):
    """
    Mark past-day records that have check-in but no check-out as absent.
    Called on dashboard, attendance, and salary views so data stays accurate
    without a background scheduler.
    Leave records and records that already have an approved correction are skipped.
    """
    today = timezone.now().date()
    qs = AttendanceRecord.objects.filter(
        check_in_time__isnull=False,
        check_out_time__isnull=True,
        date__lt=today,
    ).exclude(status="leave")
    if employee is not None:
        qs = qs.filter(employee=employee)

    # Don't flip records that an admin already approved via a correction request
    approved_pairs = list(
        AttendanceCorrectionRequest.objects.filter(status="approved", date__lt=today).values_list("employee_id", "date")
    )
    if approved_pairs:
        skip_q = Q()
        for emp_id, dt in approved_pairs:
            skip_q |= Q(employee_id=emp_id, date=dt)
        qs = qs.exclude(skip_q)

    # Also close any open breaks on past records before marking absent
    for record in qs:
        open_break = record.breaks.filter(pause_end__isnull=True).first()
        if open_break:
            open_break.pause_end = datetime.time(23, 59, 59)
            open_break.duration_minutes = int(
                (
                    datetime.datetime.combine(record.date, open_break.pause_end)
                    - datetime.datetime.combine(record.date, open_break.pause_start)
                ).total_seconds()
                / 60
            )
            open_break.save()
    qs.update(status="absent")


def generate_employee_id():
    import re

    # select_for_update locks the last row so two concurrent adds get different IDs
    with transaction.atomic():
        last = Employee.objects.select_for_update().order_by("id").last()
        if last:
            match = re.search(r"\d+$", last.employee_id)
            num = int(match.group()) + 1 if match else Employee.objects.count() + 1
        else:
            num = 1
        return f"CRF{num:03d}"


def get_panel_mode(request):
    """Returns current panel mode for a staff user."""
    if request.user.is_staff:
        return request.session.get("panel_mode", "admin")
    return "employee"


# ─── AUTH VIEWS ─────────────────────────────────────────────────────────────


def login_view(request):
    if request.user.is_authenticated:
        if request.user.is_staff and get_panel_mode(request) == "admin":
            return redirect("admin_dashboard")
        return redirect("dashboard")

    if request.method == "POST":
        form = LoginForm(request.POST)
        if form.is_valid():
            login_input = form.cleaned_data["email"].strip()
            password = form.cleaned_data["password"]
            user = None
            # Try email first
            try:
                user_obj = User.objects.get(email__iexact=login_input)
                user = authenticate(request, username=user_obj.username, password=password)
            except User.DoesNotExist:
                pass
            except User.MultipleObjectsReturned:
                messages.error(request, "Multiple accounts share this email. Contact admin.")
            # Fall back to username (e.g. admin account without email set)
            if user is None:
                user = authenticate(request, username=login_input, password=password)
            if user is not None:
                login(request, user)
                if user.is_staff:
                    request.session["panel_mode"] = "admin"
                    return redirect("admin_dashboard")
                return redirect("dashboard")
            else:
                messages.error(request, "Invalid email or password.")
    else:
        form = LoginForm()

    return render(request, "attendance/login.html", {"form": form})


def logout_view(request):
    logout(request)
    return redirect("login")


# ─── PANEL SWITCHER (Admin only) ────────────────────────────────────────────


@login_required
@user_passes_test(is_admin)
def switch_panel(request):
    """Lets admin switch between Admin Panel and Employee Panel."""
    current = get_panel_mode(request)

    if current == "admin":
        request.session["panel_mode"] = "employee"
        return redirect("dashboard")
    else:
        request.session["panel_mode"] = "admin"
        return redirect("admin_dashboard")


# ─── EMPLOYEE DASHBOARD ──────────────────────────────────────────────────────


@login_required
def dashboard(request):
    # Staff in admin mode → push them back to admin panel
    if request.user.is_staff and get_panel_mode(request) == "admin":
        return redirect("admin_dashboard")

    try:
        employee = request.user.employee
    except Employee.DoesNotExist:
        if request.user.is_staff:
            # Admin switched to employee panel but hasn't created their own profile yet
            return render(
                request,
                "attendance/no_employee_profile.html",
                {
                    "today": timezone.now().date(),
                },
            )
        messages.error(request, "Employee profile not found. Contact admin.")
        logout(request)
        return redirect("login")

    today = timezone.now().date()
    now_local = timezone.localtime(timezone.now())
    # Auto-mark past days with check-in but no check-out as absent
    _resolve_incomplete_checkins(employee)
    today_attendance = employee.get_today_attendance()

    # Break state for the dashboard attendance card
    active_break = None
    today_breaks = []
    att_state = "not_checked_in"
    checkin_elapsed_secs = 0
    break_elapsed_secs = 0

    if today_attendance and today_attendance.check_in_time:
        today_breaks = list(today_attendance.breaks.all())
        active_break = today_attendance.breaks.filter(pause_end__isnull=True).first()
        if today_attendance.check_out_time:
            att_state = "checked_out"
        elif active_break:
            att_state = "on_break"
        else:
            att_state = "working"

        # Use simple seconds-since-midnight arithmetic to avoid timezone-aware subtraction issues
        now_t = now_local.time()
        now_secs = now_t.hour * 3600 + now_t.minute * 60 + now_t.second

        cin = today_attendance.check_in_time
        cin_secs = cin.hour * 3600 + cin.minute * 60 + cin.second
        gross_elapsed = max(0, now_secs - cin_secs)

        if active_break:
            ps = active_break.pause_start
            ps_secs = ps.hour * 3600 + ps.minute * 60 + ps.second
            break_elapsed_secs = max(0, now_secs - ps_secs)

        already_broken = (today_attendance.break_minutes or 0) * 60
        if active_break:
            already_broken += break_elapsed_secs
        checkin_elapsed_secs = max(0, gross_elapsed - already_broken)

    # ── Build full 7-day view (today back to 6 days ago) ──────────────────
    week_start = today - datetime.timedelta(days=6)

    # Fetch attendance records for the week in one query
    attendance_map = {
        r.date: r for r in AttendanceRecord.objects.filter(employee=employee, date__range=(week_start, today))
    }

    # Fetch ALL approved leaves that overlap this week
    def _expand_leaves(qs, from_d, to_d):
        """Expand LeaveRequest querysets into a set of individual dates."""
        dates = set()
        for lv in qs:
            cur = lv.from_date
            while cur <= lv.to_date:
                if from_d <= cur <= to_d:
                    dates.add(cur)
                cur += datetime.timedelta(days=1)
        return dates

    week_leaves_qs = LeaveRequest.objects.filter(
        employee=employee,
        status="approved",
        from_date__lte=today,
        to_date__gte=week_start,
    )
    week_leave_dates = _expand_leaves(week_leaves_qs, week_start, today)

    # Build day-by-day list (newest first after reverse)
    week_days = []
    for i in range(6, -1, -1):
        d = today - datetime.timedelta(days=i)
        record = attendance_map.get(d)
        is_sunday = d.weekday() == 6
        # on_leave: approved leave AND no actual check-in record
        on_leave = d in week_leave_dates and not record
        week_days.append(
            {
                "date": d,
                "record": record,
                "is_sunday": is_sunday,
                "on_leave": on_leave,
                "is_today": d == today,
            }
        )
    week_days.reverse()  # newest first for display

    # ── Monthly summary ────────────────────────────────────────────────────
    import calendar as _cal

    month_start = today.replace(day=1)
    last_day = _cal.monthrange(today.year, today.month)[1]
    month_end = datetime.date(today.year, today.month, last_day)

    monthly_records = AttendanceRecord.objects.filter(
        employee=employee, date__month=today.month, date__year=today.year
    ).order_by("date")

    present_qs = monthly_records.filter(status="present", date__week_day__gt=1)
    half_day_qs = monthly_records.filter(status="half_day", date__week_day__gt=1)
    leave_rec_qs = monthly_records.filter(status="leave", date__week_day__gt=1)

    present_days = present_qs.count()
    half_days_cnt = half_day_qs.count()

    # All dates that have ANY attendance record this month
    record_dates = set(monthly_records.values_list("date", flat=True))

    # ALL approved leave days this month (full month range, not capped at today)
    # so future approved leaves are also reflected in the stat card
    month_leaves_qs = LeaveRequest.objects.filter(
        employee=employee,
        status="approved",
        from_date__lte=month_end,  # leave starts before month ends
        to_date__gte=month_start,  # leave ends after month starts
    )
    month_leave_dates = set(
        d for d in _expand_leaves(month_leaves_qs, month_start, month_end) if d.weekday() != 6  # exclude Sunday
    )

    # Leave days = records with status=leave  +  approved leave days with NO record
    leave_days = leave_rec_qs.count() + len(month_leave_dates - record_dates)

    # Absent = past working days (Mon–Sat) up to today with no record OR a record marked absent
    absent_rec_qs = monthly_records.filter(status="absent", date__week_day__gt=1, date__lte=today)
    absent_rec_dates = set(absent_rec_qs.values_list("date", flat=True))

    absent_list = []
    d = month_start
    while d <= today:
        if d.weekday() != 6 and d not in month_leave_dates:
            if d not in record_dates or d in absent_rec_dates:
                absent_list.append(d)
        d += datetime.timedelta(days=1)
    absent_days = len(absent_list)

    # Build a lookup for absent records (for check_in/check_out display in filtered_rows)
    absent_rec_map = {r.date: r for r in absent_rec_qs}

    # ── Stat-card filter ──────────────────────────────────────────────────
    stat_filter = request.GET.get("filter", "")
    filtered_rows = None  # list of dicts shown below stat cards

    if stat_filter == "present":
        filtered_rows = [
            {
                "date": r.date,
                "check_in": r.check_in_time,
                "check_out": r.check_out_time,
                "hours": r.total_hours,
                "status": r.status,
                "label": r.get_status_display(),
            }
            for r in present_qs.order_by("-date")
        ]
    elif stat_filter == "half_day":
        filtered_rows = [
            {
                "date": r.date,
                "check_in": r.check_in_time,
                "check_out": r.check_out_time,
                "hours": r.total_hours,
                "status": r.status,
                "label": r.get_status_display(),
            }
            for r in half_day_qs.order_by("-date")
        ]
    elif stat_filter == "leave":
        # Attendance records with status=leave
        leave_from_records = {r.date: r for r in leave_rec_qs}
        # Approved leave days — build unified list (record takes priority over request)
        all_leave_dates = sorted(month_leave_dates | set(leave_from_records.keys()), reverse=True)
        filtered_rows = []
        for ld in all_leave_dates:
            rec = leave_from_records.get(ld)
            filtered_rows.append(
                {
                    "date": ld,
                    "check_in": rec.check_in_time if rec else None,
                    "check_out": rec.check_out_time if rec else None,
                    "hours": rec.total_hours if rec else None,
                    "status": "leave",
                    "label": "On Leave",
                }
            )
    elif stat_filter == "absent":
        filtered_rows = []
        for d in sorted(absent_list, reverse=True):
            rec = absent_rec_map.get(d)
            filtered_rows.append(
                {
                    "date": d,
                    "check_in": rec.check_in_time if rec else None,
                    "check_out": rec.check_out_time if rec else None,
                    "hours": rec.total_hours if rec else None,
                    "status": "absent",
                    "label": "Absent",
                }
            )

    show_doc_reminder = not (employee.aadhaar_card and employee.pan_card)

    context = {
        "employee": employee,
        "today": today,
        "today_attendance": today_attendance,
        "week_days": week_days,
        "present_days": present_days,
        "half_days": half_days_cnt,
        "absent_days": absent_days,
        "leave_days": leave_days,
        "stat_filter": stat_filter,
        "filtered_rows": filtered_rows,
        "show_doc_reminder": show_doc_reminder,
        # Break feature
        "att_state": att_state,
        "active_break": active_break,
        "today_breaks": today_breaks,
        "checkin_elapsed_secs": checkin_elapsed_secs,
        "break_elapsed_secs": break_elapsed_secs,
    }
    return render(request, "attendance/dashboard.html", context)


# ─── DISMISS DOCUMENT REMINDER ───────────────────────────────────────────────


@login_required
def dismiss_document_reminder(request):
    from django.http import JsonResponse

    if request.method == "POST":
        try:
            request.user.employee.has_seen_document_reminder = True
            request.user.employee.save(update_fields=["has_seen_document_reminder"])
        except Employee.DoesNotExist:
            pass
    return JsonResponse({"ok": True})


# ─── ATTENDANCE ──────────────────────────────────────────────────────────────


@login_required
def check_in(request):
    try:
        employee = request.user.employee
    except Employee.DoesNotExist:
        return redirect("dashboard")

    today = timezone.now().date()
    now_time = timezone.localtime(timezone.now()).time()

    try:
        with transaction.atomic():
            attendance, created = AttendanceRecord.objects.get_or_create(
                employee=employee, date=today, defaults={"check_in_time": now_time, "status": "present"}
            )
    except IntegrityError:
        # Two simultaneous requests — fetch what was just created
        attendance = AttendanceRecord.objects.get(employee=employee, date=today)
        created = False

    if not created and attendance.check_in_time:
        messages.warning(request, f'Already checked in at {attendance.check_in_time.strftime("%I:%M %p")}')
    elif not created and not attendance.check_in_time:
        with transaction.atomic():
            attendance.check_in_time = now_time
            attendance.status = "present"
            attendance.save()
        messages.success(request, f'Check-in successful at {now_time.strftime("%I:%M %p")}')
    else:
        messages.success(request, f'Check-in successful at {now_time.strftime("%I:%M %p")}')

    return redirect("dashboard")


@login_required
def check_out(request):
    try:
        employee = request.user.employee
    except Employee.DoesNotExist:
        return redirect("dashboard")

    today = timezone.now().date()
    now_time = timezone.localtime(timezone.now()).time()

    try:
        with transaction.atomic():
            attendance = AttendanceRecord.objects.select_for_update().get(employee=employee, date=today)
            if not attendance.check_in_time:
                messages.error(request, "You have not checked in today.")
            elif attendance.check_out_time:
                messages.warning(request, f'Already checked out at {attendance.check_out_time.strftime("%I:%M %p")}')
            else:
                # Block checkout if a break is still active
                active_break = attendance.breaks.filter(pause_end__isnull=True).first()
                if active_break:
                    messages.error(request, "Please resume from break before checking out.")
                else:
                    attendance.check_out_time = now_time
                    attendance.save()
                    attendance.calculate_hours()
                    net = attendance.net_hours or attendance.total_hours
                    brk = attendance.break_minutes
                    status_label = attendance.get_status_display()
                    detail = f"Net: {net} hrs"
                    if brk:
                        detail += f" (break: {brk} min)"
                    messages.success(
                        request, f'Checked out at {now_time.strftime("%I:%M %p")}. {detail} — {status_label}'
                    )
    except AttendanceRecord.DoesNotExist:
        messages.error(request, "No check-in found for today. Please check in first.")

    return redirect("dashboard")


@login_required
def pause_attendance(request):
    if request.method != "POST":
        return redirect("dashboard")
    try:
        employee = request.user.employee
    except Employee.DoesNotExist:
        return redirect("dashboard")

    today = timezone.now().date()
    now_time = timezone.localtime(timezone.now()).time()
    try:
        attendance = AttendanceRecord.objects.get(employee=employee, date=today)
    except AttendanceRecord.DoesNotExist:
        messages.error(request, "You have not checked in today.")
        return redirect("dashboard")

    if not attendance.check_in_time or attendance.check_out_time:
        messages.error(request, "Cannot start break right now.")
        return redirect("dashboard")

    if attendance.breaks.filter(pause_end__isnull=True).exists():
        messages.warning(request, "You are already on a break.")
        return redirect("dashboard")

    AttendanceBreak.objects.create(attendance=attendance, pause_start=now_time)
    messages.success(request, f'Break started at {now_time.strftime("%I:%M %p")}.')
    return redirect("dashboard")


@login_required
def resume_attendance(request):
    if request.method != "POST":
        return redirect("dashboard")
    try:
        employee = request.user.employee
    except Employee.DoesNotExist:
        return redirect("dashboard")

    today = timezone.now().date()
    now_time = timezone.localtime(timezone.now()).time()
    try:
        attendance = AttendanceRecord.objects.get(employee=employee, date=today)
    except AttendanceRecord.DoesNotExist:
        messages.error(request, "No check-in found for today.")
        return redirect("dashboard")

    active_break = attendance.breaks.filter(pause_end__isnull=True).first()
    if not active_break:
        messages.warning(request, "You are not currently on a break.")
        return redirect("dashboard")

    pause_start_dt = datetime.datetime.combine(today, active_break.pause_start)
    pause_end_dt = datetime.datetime.combine(today, now_time)
    if pause_end_dt < pause_start_dt:
        pause_end_dt += datetime.timedelta(days=1)
    duration_mins = int((pause_end_dt - pause_start_dt).total_seconds() / 60)

    active_break.pause_end = now_time
    active_break.duration_minutes = duration_mins
    active_break.save()

    attendance.break_minutes = sum(b.duration_minutes for b in attendance.breaks.filter(pause_end__isnull=False))
    attendance.save(update_fields=["break_minutes"])

    messages.success(request, f"Break ended. Duration: {duration_mins} min. Back to work!")
    return redirect("dashboard")


@login_required
def my_attendance(request):
    try:
        employee = request.user.employee
    except Employee.DoesNotExist:
        return redirect("login")

    import calendar

    today = timezone.now().date()
    _resolve_incomplete_checkins(employee)
    month = int(request.GET.get("month", today.month))
    year = int(request.GET.get("year", today.year))
    stat_filter = request.GET.get("filter", "")

    records = AttendanceRecord.objects.filter(employee=employee, date__month=month, date__year=year).order_by("-date")

    present_qs = records.filter(status="present", date__week_day__gt=1)
    half_day_qs = records.filter(status="half_day", date__week_day__gt=1)
    leave_qs = records.filter(status="leave", date__week_day__gt=1)

    # Absent = working days (Mon–Sat) up to today (or end of month) with no record OR record marked absent
    record_dates = set(records.values_list("date", flat=True))
    month_start = datetime.date(year, month, 1)
    last_day = calendar.monthrange(year, month)[1]
    month_end = datetime.date(year, month, last_day)
    cutoff = min(month_end, today)  # don't count future days as absent
    absent_rec_qs2 = records.filter(status="absent", date__week_day__gt=1, date__lte=cutoff)
    absent_rec_dates2 = set(absent_rec_qs2.values_list("date", flat=True))
    absent_rec_map2 = {r.date: r for r in absent_rec_qs2}
    absent_dates = []
    d = month_start
    while d <= cutoff:
        if d.weekday() != 6 and (d not in record_dates or d in absent_rec_dates2):
            absent_dates.append(d)
        d += datetime.timedelta(days=1)

    # Filtered detail list
    filtered_rows = None
    if stat_filter == "present":
        filtered_rows = [
            {
                "date": r.date,
                "check_in": r.check_in_time,
                "check_out": r.check_out_time,
                "hours": r.total_hours,
                "status": "present",
                "label": r.get_status_display(),
            }
            for r in present_qs.order_by("-date")
        ]
    elif stat_filter == "half_day":
        filtered_rows = [
            {
                "date": r.date,
                "check_in": r.check_in_time,
                "check_out": r.check_out_time,
                "hours": r.total_hours,
                "status": "half_day",
                "label": r.get_status_display(),
            }
            for r in half_day_qs.order_by("-date")
        ]
    elif stat_filter == "leave":
        filtered_rows = [
            {"date": r.date, "check_in": None, "check_out": None, "hours": None, "status": "leave", "label": "On Leave"}
            for r in leave_qs.order_by("-date")
        ]
    elif stat_filter == "absent":
        filtered_rows = []
        for d in sorted(absent_dates, reverse=True):
            rec = absent_rec_map2.get(d)
            filtered_rows.append(
                {
                    "date": d,
                    "check_in": rec.check_in_time if rec else None,
                    "check_out": rec.check_out_time if rec else None,
                    "hours": rec.total_hours if rec else None,
                    "status": "absent",
                    "label": "Absent",
                }
            )

    context = {
        "employee": employee,
        "records": records,
        "month": month,
        "year": year,
        "month_name": calendar.month_name[month],
        "present": present_qs.count(),
        "half_day": half_day_qs.count(),
        "absent": len(absent_dates),
        "leave": leave_qs.count(),
        "stat_filter": stat_filter,
        "filtered_rows": filtered_rows,
        "today": today,
        "months": [(i, calendar.month_name[i]) for i in range(1, 13)],
        "years": range(2023, today.year + 1),
    }
    return render(request, "attendance/my_attendance.html", context)


# ─── EMPLOYEE PROFILE EDIT ───────────────────────────────────────────────────


@login_required
def edit_profile(request):
    try:
        employee = request.user.employee
    except Employee.DoesNotExist:
        messages.error(request, "No employee profile found.")
        return redirect("dashboard")

    if request.method == "POST":
        form = EmployeeProfileEditForm(request.POST, request.FILES)
        if form.is_valid():
            user = request.user
            user.first_name = form.cleaned_data["first_name"]
            user.last_name = form.cleaned_data["last_name"]
            user.email = form.cleaned_data["email"]
            user.save()

            employee.phone = form.cleaned_data["phone"]
            if form.cleaned_data.get("profile_photo"):
                employee.profile_photo = form.cleaned_data["profile_photo"]
            if form.cleaned_data.get("aadhaar_card"):
                employee.aadhaar_card = form.cleaned_data["aadhaar_card"]
            if form.cleaned_data.get("pan_card"):
                employee.pan_card = form.cleaned_data["pan_card"]
            try:
                employee.save()
            except OSError:
                employee.profile_photo = None
                employee.aadhaar_card = None
                employee.pan_card = None
                employee.save()

            # Handle optional inline password change
            current_pwd = request.POST.get("current_password", "").strip()
            new_pwd = request.POST.get("new_password", "").strip()
            conf_pwd = request.POST.get("confirm_password", "").strip()
            if current_pwd or new_pwd or conf_pwd:
                if not user.check_password(current_pwd):
                    messages.error(request, "Profile saved, but password not changed: current password is incorrect.")
                elif len(new_pwd) < 8:
                    messages.error(
                        request, "Profile saved, but password not changed: new password must be at least 8 characters."
                    )
                elif new_pwd != conf_pwd:
                    messages.error(request, "Profile saved, but password not changed: new passwords do not match.")
                else:
                    user.set_password(new_pwd)
                    user.save()
                    update_session_auth_hash(request, user)
                    messages.success(request, "Profile and password updated successfully!")
                    return redirect("edit_profile")

            messages.success(request, "Profile updated successfully!")
            return redirect("edit_profile")
    else:
        form = EmployeeProfileEditForm(
            initial={
                "first_name": request.user.first_name,
                "last_name": request.user.last_name,
                "email": request.user.email,
                "phone": employee.phone,
            }
        )

    return render(
        request,
        "attendance/edit_profile.html",
        {
            "form": form,
            "employee": employee,
        },
    )


@login_required
def change_password(request):
    if request.method != "POST":
        return redirect("edit_profile")
    current = request.POST.get("current_password", "")
    new1 = request.POST.get("new_password", "")
    new2 = request.POST.get("confirm_password", "")
    if not request.user.check_password(current):
        messages.error(request, "Current password is incorrect.")
        return redirect("edit_profile")
    if len(new1) < 8:
        messages.error(request, "New password must be at least 8 characters.")
        return redirect("edit_profile")
    if new1 != new2:
        messages.error(request, "New passwords do not match.")
        return redirect("edit_profile")
    request.user.set_password(new1)
    request.user.save()
    update_session_auth_hash(request, request.user)
    messages.success(request, "Password changed successfully!")
    return redirect("edit_profile")


# ─── ADMIN DASHBOARD ─────────────────────────────────────────────────────────


@login_required
@user_passes_test(is_admin)
def admin_dashboard(request):
    # If admin switched to employee mode, redirect them to employee panel
    if get_panel_mode(request) == "employee":
        return redirect("dashboard")

    today = timezone.now().date()
    status_filter = request.GET.get("filter", "")  # 'present' | 'absent' | 'in_office'

    all_active = Employee.objects.filter(is_active=True).select_related("user", "department")
    total_employees = all_active.count()

    present_records = AttendanceRecord.objects.filter(date=today, status__in=["present", "half_day"]).select_related(
        "employee__user", "employee__department"
    )

    present_emp_ids = present_records.values_list("employee_id", flat=True)
    today_present = present_records.count()

    absent_employees = all_active.exclude(id__in=present_emp_ids)
    today_absent = absent_employees.count()

    checked_in_now = AttendanceRecord.objects.filter(
        date=today, check_in_time__isnull=False, check_out_time__isnull=True
    ).select_related("employee__user", "employee__department")

    # Split into Working vs On Break using AttendanceBreak
    on_break_record_ids = set(
        AttendanceBreak.objects.filter(attendance__date=today, pause_end__isnull=True).values_list(
            "attendance_id", flat=True
        )
    )
    working_records = [r for r in checked_in_now if r.id not in on_break_record_ids]
    on_break_records = [r for r in checked_in_now if r.id in on_break_record_ids]

    recent_activity = (
        AttendanceRecord.objects.filter(date=today).select_related("employee__user").order_by("-check_in_time")[:10]
    )

    # Filtered list for stat card clicks
    filtered_label = ""
    filtered_list = None
    if status_filter == "present":
        filtered_list = present_records.order_by("-check_in_time")
        filtered_label = "Present Today"
    elif status_filter == "absent":
        filtered_list = absent_employees.order_by("user__first_name")
        filtered_label = "Absent Today"
    elif status_filter == "in_office":
        filtered_list = checked_in_now.order_by("-check_in_time")
        filtered_label = "Currently In Office"

    context = {
        "total_employees": total_employees,
        "today_present": today_present,
        "today_absent": today_absent,
        "checked_in_now": checked_in_now,
        "working_records": working_records,
        "on_break_records": on_break_records,
        "recent_activity": recent_activity,
        "today": today,
        "status_filter": status_filter,
        "filtered_list": filtered_list,
        "filtered_label": filtered_label,
    }
    return render(request, "attendance/admin_dashboard.html", context)


# ─── ADMIN EMPLOYEE MANAGEMENT ───────────────────────────────────────────────


@login_required
@user_passes_test(is_admin)
def manage_employees(request):
    departments = Department.objects.all()
    search = request.GET.get("search", "")
    dept_filter = request.GET.get("department", "")
    tab = request.GET.get("tab", "active")  # "active" | "fired"

    base_qs = Employee.objects.select_related("user", "department")
    active_qs = base_qs.filter(is_active=True)
    fired_qs = base_qs.filter(is_active=False)

    if search:
        q = (
            Q(user__first_name__icontains=search)
            | Q(user__last_name__icontains=search)
            | Q(employee_id__icontains=search)
            | Q(designation__icontains=search)
        )
        active_qs = active_qs.filter(q)
        fired_qs = fired_qs.filter(q)
    if dept_filter:
        active_qs = active_qs.filter(department__id=dept_filter)
        fired_qs = fired_qs.filter(department__id=dept_filter)

    employees = fired_qs if tab == "fired" else active_qs

    return render(
        request,
        "attendance/employees.html",
        {
            "employees": employees,
            "departments": departments,
            "search": search,
            "tab": tab,
            "active_count": active_qs.count(),
            "fired_count": fired_qs.count(),
        },
    )


@login_required
@user_passes_test(is_admin)
def view_employee_detail(request, emp_id):
    employee = get_object_or_404(Employee, id=emp_id)
    recent_attendance = AttendanceRecord.objects.filter(employee=employee).order_by("-date")[:10]
    recent_leaves = LeaveRequest.objects.filter(employee=employee)[:5]
    recent_tickets = SupportTicket.objects.filter(employee=employee)[:5]

    return render(
        request,
        "attendance/employee_detail.html",
        {
            "employee": employee,
            "recent_attendance": recent_attendance,
            "recent_leaves": recent_leaves,
            "recent_tickets": recent_tickets,
        },
    )


@login_required
@user_passes_test(is_admin)
def add_employee(request):
    next_id = generate_employee_id()
    if request.method == "POST":
        user_form = UserForm(request.POST)
        emp_form = EmployeeForm(request.POST, request.FILES)
        if user_form.is_valid() and emp_form.is_valid():
            try:
                with transaction.atomic():
                    user = user_form.save(commit=False)
                    user.set_password(user_form.cleaned_data["password"])
                    user.save()
                    employee = emp_form.save(commit=False)
                    employee.user = user
                    employee.employee_id = next_id
                    try:
                        employee.save()
                    except OSError:
                        # File system not writable (e.g. Vercel) — save without photo
                        employee.profile_photo = None
                        employee.save()
            except IntegrityError:
                messages.error(request, "A duplicate was detected. Please try again.")
                return redirect("add_employee")
            messages.success(request, f"Employee {user.get_full_name()} added! Employee ID: {next_id}")
            return redirect("manage_employees")
        else:
            messages.error(request, "Please fix the errors below.")
    else:
        user_form = UserForm()
        emp_form = EmployeeForm()

    return render(
        request,
        "attendance/add_employee.html",
        {
            "user_form": user_form,
            "emp_form": emp_form,
            "next_employee_id": next_id,
        },
    )


# ─── EMPLOYEE EDIT / ACTIONS (Admin) ─────────────────────────────────────────


@login_required
@user_passes_test(is_admin)
def edit_employee_admin(request, emp_id):
    employee = get_object_or_404(Employee, id=emp_id)
    user = employee.user

    if request.method == "POST":
        form = AdminEmployeeEditForm(request.POST, request.FILES)
        if form.is_valid():
            user.first_name = form.cleaned_data["first_name"]
            user.last_name = form.cleaned_data["last_name"]
            user.email = form.cleaned_data["email"]
            user.username = form.cleaned_data["username"]
            user.is_active = form.cleaned_data.get("is_active_user", False)
            user.save()

            employee.employee_id = form.cleaned_data["employee_id"]
            employee.department = form.cleaned_data["department"]
            employee.designation = form.cleaned_data["designation"]
            employee.phone = form.cleaned_data["phone"]
            employee.date_joined = form.cleaned_data["date_joined"]
            if form.cleaned_data.get("profile_photo"):
                employee.profile_photo = form.cleaned_data["profile_photo"]
            if form.cleaned_data.get("aadhaar_card"):
                employee.aadhaar_card = form.cleaned_data["aadhaar_card"]
            if form.cleaned_data.get("pan_card"):
                employee.pan_card = form.cleaned_data["pan_card"]
            try:
                employee.save()
            except OSError:
                employee.profile_photo = None
                employee.aadhaar_card = None
                employee.pan_card = None
                employee.save()

            messages.success(request, f"Employee {user.get_full_name()} updated successfully!")
            return redirect("employee_detail", emp_id=emp_id)
    else:
        form = AdminEmployeeEditForm(
            initial={
                "first_name": user.first_name,
                "last_name": user.last_name,
                "email": user.email,
                "username": user.username,
                "is_active_user": user.is_active,
                "employee_id": employee.employee_id,
                "department": employee.department,
                "designation": employee.designation,
                "phone": employee.phone,
                "date_joined": employee.date_joined,
            }
        )

    return render(
        request,
        "attendance/edit_employee_admin.html",
        {
            "form": form,
            "employee": employee,
        },
    )


@login_required
@user_passes_test(is_admin)
def fire_employee(request, emp_id):
    employee = get_object_or_404(Employee, id=emp_id)
    if request.method != "POST":
        return redirect("employee_detail", emp_id=emp_id)
    if employee.user == request.user:
        messages.error(request, "You cannot fire yourself.")
        return redirect("employee_detail", emp_id=emp_id)
    employee.is_active = False
    employee.save()
    employee.user.is_active = False
    employee.user.save()
    messages.success(
        request, f"{employee.user.get_full_name()} has been terminated. Their login access has been removed."
    )
    return redirect("manage_employees")


@login_required
@user_passes_test(is_admin)
def rehire_employee(request, emp_id):
    employee = get_object_or_404(Employee, id=emp_id)
    if request.method != "POST":
        return redirect("employee_detail", emp_id=emp_id)
    employee.is_active = True
    employee.save()
    employee.user.is_active = True
    employee.user.save()
    messages.success(
        request, f"{employee.user.get_full_name()} has been rehired. Their login access has been restored."
    )
    return redirect("employee_detail", emp_id=emp_id)


@login_required
@user_passes_test(is_admin)
def toggle_employee_active(request, emp_id):
    employee = get_object_or_404(Employee, id=emp_id)
    if request.method == "POST":
        employee.is_active = not employee.is_active
        employee.save()
        status = "activated" if employee.is_active else "deactivated"
        messages.success(request, f"{employee.user.get_full_name()} has been {status}.")
    return redirect("employee_detail", emp_id=emp_id)


@login_required
@user_passes_test(is_admin)
def toggle_employee_admin(request, emp_id):
    employee = get_object_or_404(Employee, id=emp_id)
    if request.method == "POST":
        user = employee.user
        if user == request.user:
            messages.error(request, "You cannot change your own admin status.")
            return redirect("employee_detail", emp_id=emp_id)
        if not user.email:
            messages.error(request, f"Set an email for {user.get_full_name()} first — they need it to log in.")
            return redirect("edit_employee_admin", emp_id=emp_id)
        user.is_staff = not user.is_staff
        user.save()
        action = "granted admin access" if user.is_staff else "revoked admin access"
        messages.success(request, f"{user.get_full_name()} has been {action}.")
    return redirect("employee_detail", emp_id=emp_id)


@login_required
@user_passes_test(is_admin)
def reset_employee_password(request, emp_id):
    employee = get_object_or_404(Employee, id=emp_id)
    if request.method == "POST":
        form = AdminPasswordChangeForm(request.POST)
        if form.is_valid():
            employee.user.set_password(form.cleaned_data["new_password"])
            employee.user.save()
            messages.success(request, f"Password for {employee.user.get_full_name()} reset successfully!")
            return redirect("employee_detail", emp_id=emp_id)
    else:
        form = AdminPasswordChangeForm()

    return render(
        request,
        "attendance/reset_employee_password.html",
        {
            "form": form,
            "employee": employee,
        },
    )


# ─── ATTENDANCE RECORD MANAGEMENT (Admin) ────────────────────────────────────


@login_required
@user_passes_test(is_admin)
def admin_attendance_records(request):
    # Resolve all employees' incomplete past-day check-ins before displaying
    _resolve_incomplete_checkins()
    records = AttendanceRecord.objects.select_related("employee__user", "employee__department").order_by(
        "-date", "-check_in_time"
    )

    emp_filter = request.GET.get("employee", "")
    date_filter = request.GET.get("date", "")
    status_filter = request.GET.get("status", "")

    if emp_filter:
        records = records.filter(employee__id=emp_filter)
    if date_filter:
        try:
            records = records.filter(date=datetime.date.fromisoformat(date_filter))
        except ValueError:
            pass
    if status_filter:
        records = records.filter(status=status_filter)

    employees = Employee.objects.filter(is_active=True).select_related("user")
    total = records.count()

    return render(
        request,
        "attendance/admin_attendance_records.html",
        {
            "records": records[:200],
            "employees": employees,
            "emp_filter": emp_filter,
            "date_filter": date_filter,
            "status_filter": status_filter,
            "total": total,
            "status_choices": AttendanceRecord.STATUS_CHOICES,
        },
    )


@login_required
@user_passes_test(is_admin)
def add_attendance_record(request):
    if request.method == "POST":
        form = AttendanceRecordForm(request.POST)
        if form.is_valid():
            record = form.save(commit=False)
            # If check-in exists, ensure status is at least 'present' before recalc
            if record.check_in_time and not record.check_out_time:
                record.status = "present"
                record.save()
            else:
                record.save()
                record.calculate_hours()  # auto-sets status based on hours
            messages.success(request, "Attendance record added.")
            return redirect("admin_attendance_records")
    else:
        form = AttendanceRecordForm(initial={"date": timezone.now().date()})

    return render(
        request,
        "attendance/add_edit_attendance.html",
        {
            "form": form,
            "title": "Add Attendance Record",
            "is_edit": False,
        },
    )


@login_required
@user_passes_test(is_admin)
def edit_attendance_record(request, record_id):
    record = get_object_or_404(AttendanceRecord, id=record_id)
    if request.method == "POST":
        form = AttendanceRecordForm(request.POST, instance=record)
        if form.is_valid():
            saved = form.save(commit=False)
            if saved.check_in_time and not saved.check_out_time:
                # Check-in only → present (no recalc needed, just ensure status)
                saved.status = "present"
                saved.total_hours = None
                saved.save()
            else:
                saved.save()
                saved.calculate_hours()  # auto-sets status based on hours
            messages.success(request, "Attendance record updated.")
            return redirect("admin_attendance_records")
    else:
        form = AttendanceRecordForm(instance=record)

    return render(
        request,
        "attendance/add_edit_attendance.html",
        {
            "form": form,
            "title": "Edit Attendance Record",
            "is_edit": True,
            "record": record,
        },
    )


@login_required
@user_passes_test(is_admin)
def delete_attendance_record(request, record_id):
    record = get_object_or_404(AttendanceRecord, id=record_id)
    if request.method == "POST":
        record.delete()
        messages.success(request, "Attendance record deleted.")
    return redirect("admin_attendance_records")


# ─── DEPARTMENT MANAGEMENT ───────────────────────────────────────────────────


@login_required
@user_passes_test(is_admin)
def manage_departments(request):
    departments = Department.objects.annotate(employee_count=Count("employee")).order_by("name")

    if request.method == "POST":
        form = DepartmentForm(request.POST)
        if form.is_valid():
            dept = form.save()
            messages.success(request, f'Department "{dept.name}" added successfully!')
            return redirect("manage_departments")
    else:
        form = DepartmentForm()

    return render(
        request,
        "attendance/manage_departments.html",
        {
            "departments": departments,
            "form": form,
        },
    )


@login_required
@user_passes_test(is_admin)
def edit_department(request, dept_id):
    department = get_object_or_404(Department, id=dept_id)

    if request.method == "POST":
        form = DepartmentForm(request.POST, instance=department)
        if form.is_valid():
            form.save()
            messages.success(request, f'Department "{department.name}" updated!')
            return redirect("manage_departments")
    else:
        form = DepartmentForm(instance=department)

    return render(
        request,
        "attendance/edit_department.html",
        {
            "form": form,
            "department": department,
            "employee_count": department.employee_set.count(),
        },
    )


@login_required
@user_passes_test(is_admin)
def delete_department(request, dept_id):
    department = get_object_or_404(Department, id=dept_id)
    if request.method == "POST":
        name = department.name
        department.delete()
        messages.success(request, f'Department "{name}" deleted.')
    return redirect("manage_departments")


# ─── REPORTS ─────────────────────────────────────────────────────────────────


@login_required
@user_passes_test(is_admin)
def reports(request):
    import calendar

    month = int(request.GET.get("month", timezone.now().month))
    year = int(request.GET.get("year", timezone.now().year))
    employees = Employee.objects.filter(is_active=True).select_related("user")

    report_data = []
    for emp in employees:
        records = AttendanceRecord.objects.filter(employee=emp, date__month=month, date__year=year)
        total_hours = sum([float(r.total_hours or 0) for r in records])
        report_data.append(
            {
                "employee": emp,
                "present": records.filter(status="present", date__week_day__gt=1).count(),
                "half_day": records.filter(status="half_day", date__week_day__gt=1).count(),
                "absent": records.filter(status="absent", date__week_day__gt=1).count(),
                "leave": records.filter(status="leave", date__week_day__gt=1).count(),
                "total_hours": round(total_hours, 2),
            }
        )

    return render(
        request,
        "attendance/reports.html",
        {
            "report_data": report_data,
            "month": month,
            "year": year,
            "month_name": calendar.month_name[month],
            "months": [(i, calendar.month_name[i]) for i in range(1, 13)],
            "years": range(2023, timezone.now().year + 1),
        },
    )


# ─── EXCEL EXPORT ────────────────────────────────────────────────────────────


@login_required
@user_passes_test(is_admin)
def export_monthly_report_excel(request):
    import calendar

    import openpyxl
    from django.http import HttpResponse
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    month = int(request.GET.get("month", timezone.now().month))
    year = int(request.GET.get("year", timezone.now().year))
    month_name_str = calendar.month_name[month]

    employees = Employee.objects.filter(is_active=True).select_related("user", "department")

    # Bulk-fetch salary records to avoid N+1 queries
    salary_map = {s.employee_id: s for s in SalaryRecord.objects.filter(month=month, year=year)}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{month_name_str[:3]} {year}"

    # Title row — spans 15 columns (A:O)
    ws.merge_cells("A1:O1")
    ws["A1"] = f"Crefio | Monthly Attendance & Salary Report: {month_name_str} {year}"
    ws["A1"].font = Font(bold=True, size=13, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="111827", end_color="111827", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # Header row
    headers = [
        "Emp ID",
        "Full Name",
        "Department",
        "Designation",  # 1-4
        "Present",
        "Half Day",
        "Absent",
        "Leave",
        "Total Hours",  # 5-9
        "Attendance %",  # 10
        "Basic Salary (₹)",
        "Deductions (₹)",
        "Net Salary (₹)",  # 11-13
        "Payment Status",
        "Paid Date",  # 14-15
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = Font(bold=True, color="111827", size=10)
        if col >= 11:  # salary columns — teal header
            cell.fill = PatternFill(start_color="A7F3D0", end_color="A7F3D0", fill_type="solid")
        else:
            cell.fill = PatternFill(start_color="BAF2BF", end_color="BAF2BF", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 22

    # Data rows
    for row_idx, emp in enumerate(employees, 3):
        records = AttendanceRecord.objects.filter(employee=emp, date__month=month, date__year=year)
        present = records.filter(status="present", date__week_day__gt=1).count()
        half_day = records.filter(status="half_day", date__week_day__gt=1).count()
        absent = records.filter(status="absent", date__week_day__gt=1).count()
        leave = records.filter(status="leave", date__week_day__gt=1).count()
        total_hrs = round(sum(float(r.total_hours or 0) for r in records), 2)
        total = present + half_day + absent + leave
        pct = f"{round(present / total * 100, 1)}%" if total > 0 else "N/A"

        salary = salary_map.get(emp.id)
        basic_salary = float(salary.basic_salary) if salary else None
        deductions = float(salary.deductions) if salary else None
        net_salary = float(salary.net_salary) if salary else None
        payment_status = "Paid" if salary and salary.is_paid else "Pending" if salary else "Not Set"
        paid_date = salary.paid_date.strftime("%d %b %Y") if salary and salary.is_paid and salary.paid_date else "—"

        row_data = [
            emp.employee_id,
            emp.user.get_full_name() or emp.user.username,
            emp.department.name if emp.department else "-",
            emp.designation,
            present,
            half_day,
            absent,
            leave,
            total_hrs,
            pct,
            basic_salary if basic_salary is not None else "Not Set",
            deductions if deductions is not None else "Not Set",
            net_salary if net_salary is not None else "Not Set",
            payment_status,
            paid_date,
        ]

        fill_color = "FFFFFF" if row_idx % 2 == 0 else "F7F8FA"
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            cell.alignment = Alignment(
                horizontal="center" if col > 4 else "left",
                vertical="center",
            )

            if col in (11, 12, 13) and isinstance(val, float):  # money columns
                cell.number_format = "₹#,##0.00"
                cell.font = Font(bold=(col == 13), size=10)  # net salary bold

            elif col == 14:  # Payment Status
                if payment_status == "Paid":
                    cell.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
                    cell.font = Font(bold=True, color="065F46", size=10)
                elif payment_status == "Pending":
                    cell.fill = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid")
                    cell.font = Font(bold=True, color="854D0E", size=10)
                else:
                    cell.font = Font(color="6B7280", size=10)

    # Column widths
    col_widths = [13, 24, 18, 18, 10, 10, 10, 10, 14, 13, 16, 16, 16, 15, 14]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="Crefio_Attendance_{month_name_str}_{year}.xlsx"'
    wb.save(response)
    return response


@login_required
@user_passes_test(is_admin)
def export_employee_monthly_excel(request, emp_id):
    """Download a single employee's day-by-day attendance for the selected month as Excel."""
    import calendar

    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    employee = get_object_or_404(Employee, id=emp_id)
    month = int(request.GET.get("month", timezone.now().month))
    year = int(request.GET.get("year", timezone.now().year))
    month_name_str = calendar.month_name[month]

    records_qs = AttendanceRecord.objects.filter(employee=employee, date__month=month, date__year=year)
    records_map = {r.date: r for r in records_qs}

    num_days = calendar.monthrange(year, month)[1]
    today = timezone.now().date()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{employee.employee_id}_{month_name_str[:3]}{year}"

    # Title
    ws.merge_cells("A1:H1")
    ws["A1"] = (
        f"Crefio | {employee.user.get_full_name()} ({employee.employee_id}) | Attendance: {month_name_str} {year}"
    )
    ws["A1"].font = Font(bold=True, size=12, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="111827", end_color="111827", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    headers = ["Date", "Day", "Check-In", "Check-Out", "Hours Worked", "Status", "Notes"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = Font(bold=True, size=10)
        cell.fill = PatternFill(start_color="BAF2BF", end_color="BAF2BF", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    status_colors = {
        "present": "D1FAE5",
        "half_day": "DBEAFE",
        "absent": "FEE2E2",
        "leave": "FEF9C3",
    }

    for day in range(1, num_days + 1):
        d = datetime.date(year, month, day)
        record = records_map.get(d)
        is_sunday = d.weekday() == 6
        is_future = d > today
        row = day + 2

        fill_color = "F0F0F0" if is_sunday else "FAFAFA" if is_future else "FFFFFF"
        if record and not is_sunday and not is_future:
            fill_color = status_colors.get(record.status, "FFFFFF")

        check_in = record.check_in_time.strftime("%I:%M %p") if record and record.check_in_time else "—"
        check_out = record.check_out_time.strftime("%I:%M %p") if record and record.check_out_time else "—"
        hours = float(record.total_hours) if record and record.total_hours else "—"

        if is_sunday:
            status_label = "Sunday / Holiday"
        elif is_future:
            status_label = "Future"
        elif record:
            status_label = record.get_status_display()
        else:
            status_label = "Absent / No Record"

        row_data = [
            d.strftime("%d %b %Y"),
            d.strftime("%A"),
            check_in,
            check_out,
            hours,
            status_label,
            record.notes if record else "",
        ]

        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            cell.alignment = Alignment(horizontal="center" if col > 2 else "left", vertical="center")

    col_widths = [14, 14, 12, 12, 14, 18, 30]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    fname = f"Crefio_{employee.employee_id}_{month_name_str}_{year}.xlsx"
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{fname}"'
    wb.save(response)
    return response


@login_required
@user_passes_test(is_admin)
def edit_day_attendance(request, emp_id):
    """Create or update a single day's attendance record from the monthly detail page."""
    employee = get_object_or_404(Employee, id=emp_id)
    if request.method != "POST":
        return redirect("employee_monthly_detail", emp_id=emp_id)

    date_str = request.POST.get("date")
    try:
        record_date = datetime.date.fromisoformat(date_str)
    except (ValueError, TypeError):
        messages.error(request, "Invalid date.")
        return redirect("employee_monthly_detail", emp_id=emp_id)

    check_in_raw = request.POST.get("check_in_time", "").strip()
    check_out_raw = request.POST.get("check_out_time", "").strip()
    status_override = request.POST.get("status", "").strip()
    notes = request.POST.get("notes", "").strip()

    # Parse HH:MM strings into proper time objects so calculate_hours() doesn't crash
    def parse_time(t):
        try:
            return datetime.time.fromisoformat(t) if t else None
        except ValueError:
            return None

    check_in_time = parse_time(check_in_raw)
    check_out_time = parse_time(check_out_raw)

    record, _ = AttendanceRecord.objects.get_or_create(
        employee=employee, date=record_date, defaults={"status": "absent"}
    )

    record.check_in_time = check_in_time
    record.check_out_time = check_out_time
    record.notes = notes

    if check_in_time and check_out_time:
        # Both times present — calculate hours, then honour status override if set
        record.save()
        record.calculate_hours()
        if status_override in ("present", "absent", "half_day", "leave") and status_override != record.status:
            record.status = status_override
            record.save()
    elif check_in_time:
        record.status = status_override if status_override in ("present", "absent", "half_day", "leave") else "present"
        record.total_hours = None
        record.save()
    else:
        record.status = status_override if status_override in ("present", "absent", "half_day", "leave") else "absent"
        record.total_hours = None
        record.save()

    messages.success(request, f"Attendance for {record_date.strftime('%d %b %Y')} updated.")
    from django.urls import reverse

    url = reverse("employee_monthly_detail", args=[emp_id])
    return redirect(f"{url}?month={record_date.month}&year={record_date.year}")


# ─── DETAILED REPORT VIEWS ───────────────────────────────────────────────────


@login_required
@user_passes_test(is_admin)
def employee_monthly_detail(request, emp_id):
    """
    Full month breakdown for one employee:
    every day → check-in time, check-out time, hours, status.
    """
    import calendar

    employee = get_object_or_404(Employee, id=emp_id)
    month = int(request.GET.get("month", timezone.now().month))
    year = int(request.GET.get("year", timezone.now().year))

    _resolve_incomplete_checkins(employee)
    # All records this employee has for the chosen month
    records_qs = AttendanceRecord.objects.filter(employee=employee, date__month=month, date__year=year)
    records_map = {r.date: r for r in records_qs}

    # Build a row for every calendar day in the month
    num_days = calendar.monthrange(year, month)[1]
    today = timezone.now().date()
    daily_rows = []

    for day in range(1, num_days + 1):
        d = datetime.date(year, month, day)
        record = records_map.get(d)
        weekday_name = d.strftime("%A")
        is_sunday = d.weekday() == 6  # Sunday only — weekly holiday
        is_future = d > today

        daily_rows.append(
            {
                "date": d,
                "day_name": weekday_name,
                "is_weekend": is_sunday,
                "is_future": is_future,
                "record": record,
            }
        )

    # Days with no record, not Sunday, not future → count as absent
    no_record_absent = sum(
        1 for row in daily_rows if not row["record"] and not row["is_weekend"] and not row["is_future"]
    )
    present = records_qs.filter(status="present", date__week_day__gt=1).count()
    half_day = records_qs.filter(status="half_day", date__week_day__gt=1).count()
    absent = records_qs.filter(status="absent", date__week_day__gt=1).count() + no_record_absent
    leave = records_qs.filter(status="leave", date__week_day__gt=1).count()
    total_hours = sum(float(r.total_hours or 0) for r in records_qs)
    avg_checkin = None
    avg_checkout = None

    checkin_times = [r.check_in_time for r in records_qs if r.check_in_time]
    checkout_times = [r.check_out_time for r in records_qs if r.check_out_time]

    if checkin_times:
        avg_sec = sum(t.hour * 3600 + t.minute * 60 + t.second for t in checkin_times) // len(checkin_times)
        avg_checkin = f"{avg_sec // 3600:02d}:{(avg_sec % 3600) // 60:02d}"
    if checkout_times:
        avg_sec = sum(t.hour * 3600 + t.minute * 60 + t.second for t in checkout_times) // len(checkout_times)
        avg_checkout = f"{avg_sec // 3600:02d}:{(avg_sec % 3600) // 60:02d}"

    return render(
        request,
        "attendance/employee_monthly_detail.html",
        {
            "employee": employee,
            "daily_rows": daily_rows,
            "month": month,
            "year": year,
            "month_name": calendar.month_name[month],
            "present": present,
            "half_day": half_day,
            "absent": absent,
            "leave": leave,
            "total_hours": round(total_hours, 2),
            "avg_checkin": avg_checkin,
            "avg_checkout": avg_checkout,
            "months": [(i, calendar.month_name[i]) for i in range(1, 13)],
            "years": range(2023, timezone.now().year + 1),
        },
    )


@login_required
@user_passes_test(is_admin)
def date_wise_report(request):
    """
    Pick any date → see every employee's check-in / check-out for that day.
    Optional filter: ?filter=present | absent | total (default)
    """
    date_str = request.GET.get("date", str(timezone.now().date()))
    try:
        selected_date = datetime.date.fromisoformat(date_str)
    except ValueError:
        selected_date = timezone.now().date()

    status_filter = request.GET.get("filter", "total")

    all_employees = Employee.objects.filter(is_active=True).select_related("user", "department")
    records_map = {
        r.employee_id: r for r in AttendanceRecord.objects.filter(date=selected_date).select_related("employee__user")
    }

    all_rows = []
    for emp in all_employees:
        record = records_map.get(emp.id)
        all_rows.append({"employee": emp, "record": record})

    present_count = sum(1 for r in all_rows if r["record"] and r["record"].check_in_time)
    absent_count = len(all_rows) - present_count

    # Apply filter
    if status_filter == "present":
        rows = [r for r in all_rows if r["record"] and r["record"].check_in_time]
    elif status_filter == "absent":
        rows = [r for r in all_rows if not r["record"] or not r["record"].check_in_time]
    else:
        rows = all_rows

    return render(
        request,
        "attendance/date_wise_report.html",
        {
            "rows": rows,
            "selected_date": selected_date,
            "present_count": present_count,
            "absent_count": absent_count,
            "total": len(all_rows),
            "status_filter": status_filter,
        },
    )


# ─── LEAVE REQUEST VIEWS ──────────────────────────────────────────────────────


@login_required
def apply_leave(request):
    from django.http import JsonResponse

    try:
        employee = request.user.employee
    except Employee.DoesNotExist:
        return redirect("dashboard")

    if request.method == "POST":
        form = LeaveRequestForm(request.POST)
        is_ajax = request.headers.get("X-Requested-With") == "XMLHttpRequest"
        if form.is_valid():
            leave = form.save(commit=False)
            leave.employee = employee
            leave.save()
            email_service.notify_admin_leave_applied(leave)
            if is_ajax:
                all_leaves = LeaveRequest.objects.filter(employee=employee)
                return JsonResponse(
                    {
                        "success": True,
                        "leave": {
                            "id": leave.id,
                            "leave_id": f"LV{leave.id:04d}",
                            "type_display": leave.get_leave_type_display(),
                            "from_date": leave.from_date.strftime("%d %b %Y"),
                            "to_date": leave.to_date.strftime("%d %b %Y"),
                            "total_days": leave.total_days,
                            "reason": leave.reason[:120],
                            "status": leave.status,
                            "applied_on": leave.applied_on.strftime("%d %b %Y"),
                        },
                        "counts": {
                            "pending": all_leaves.filter(status="pending").count(),
                            "approved": all_leaves.filter(status="approved").count(),
                            "rejected": all_leaves.filter(status="rejected").count(),
                        },
                    }
                )
            messages.success(request, "Leave request submitted! Admin will review it soon.")
            return redirect("my_leaves")
        else:
            if is_ajax:
                return JsonResponse({"success": False, "errors": form.errors})
    else:
        form = LeaveRequestForm()

    return render(request, "attendance/apply_leave.html", {"form": form, "employee": employee})


@login_required
def my_leaves(request):
    try:
        employee = request.user.employee
    except Employee.DoesNotExist:
        return redirect("dashboard")

    filter_status = request.GET.get("filter", "")
    search_q = request.GET.get("search", "").strip()

    all_leaves = LeaveRequest.objects.filter(employee=employee)
    leaves = all_leaves.order_by("-applied_on")

    if filter_status in ("pending", "approved", "rejected"):
        leaves = leaves.filter(status=filter_status)

    if search_q:
        leaves = leaves.filter(Q(leave_type__icontains=search_q) | Q(reason__icontains=search_q))

    form = LeaveRequestForm()  # for the inline modal
    return render(
        request,
        "attendance/my_leaves.html",
        {
            "leaves": leaves,
            "pending": all_leaves.filter(status="pending").count(),
            "approved": all_leaves.filter(status="approved").count(),
            "rejected": all_leaves.filter(status="rejected").count(),
            "filter_status": filter_status,
            "search_q": search_q,
            "form": form,
            "employee": employee,
            "apply_url": "/leave/apply/",
        },
    )


# ─── SUPPORT TICKET VIEWS ────────────────────────────────────────────────────


@login_required
def raise_ticket(request):
    try:
        employee = request.user.employee
    except Employee.DoesNotExist:
        return redirect("dashboard")

    is_ajax = request.headers.get("X-Requested-With") == "XMLHttpRequest"

    if request.method == "POST":
        form = SupportTicketForm(request.POST)
        if form.is_valid():
            ticket = form.save(commit=False)
            ticket.employee = employee
            ticket.save()
            email_service.notify_admin_ticket_raised(ticket)
            if is_ajax:
                all_tickets = SupportTicket.objects.filter(employee=employee)
                open_count = all_tickets.filter(status__in=["open", "in_progress", "waiting"]).count()
                resolved_count = all_tickets.filter(status__in=["resolved", "closed"]).count()
                return JsonResponse(
                    {
                        "success": True,
                        "ticket": {
                            "id": ticket.id,
                            "ticket_id": f"TK{ticket.id:04d}",
                            "subject": ticket.subject,
                            "category_display": ticket.get_category_display(),
                            "priority": ticket.priority,
                            "priority_display": ticket.get_priority_display(),
                            "status": ticket.status,
                            "created_at": ticket.created_at.strftime("%d %b %Y"),
                        },
                        "counts": {
                            "open_count": open_count,
                            "resolved_count": resolved_count,
                        },
                    }
                )
            messages.success(request, f"Ticket #{ticket.id} raised successfully! Admin will respond soon.")
            return redirect("my_tickets")
        else:
            if is_ajax:
                return JsonResponse({"success": False, "errors": form.errors})
    else:
        form = SupportTicketForm()

    return render(request, "attendance/raise_ticket.html", {"form": form, "employee": employee})


@login_required
def edit_ticket(request, ticket_id):
    try:
        employee = request.user.employee
    except Employee.DoesNotExist:
        return redirect("dashboard")

    ticket = get_object_or_404(SupportTicket, id=ticket_id, employee=employee)

    if ticket.status not in ("open", "in_progress"):
        messages.error(request, "This ticket can no longer be edited.")
        return redirect("my_tickets")

    if request.method == "POST":
        form = SupportTicketForm(request.POST, instance=ticket)
        if form.is_valid():
            form.save()
            messages.success(request, f"Ticket #{ticket.id} updated successfully.")
            return redirect("my_tickets")
    else:
        form = SupportTicketForm(instance=ticket)

    return render(
        request,
        "attendance/edit_ticket.html",
        {
            "form": form,
            "ticket": ticket,
            "employee": employee,
        },
    )


@login_required
def my_tickets(request):
    try:
        employee = request.user.employee
    except Employee.DoesNotExist:
        return redirect("dashboard")

    filter_status = request.GET.get("filter", "")
    search_q = request.GET.get("search", "").strip()

    all_tickets = SupportTicket.objects.filter(employee=employee)
    tickets = all_tickets.order_by("-created_at")

    if filter_status == "open":
        tickets = tickets.filter(status__in=["open", "in_progress", "waiting"])
    elif filter_status == "resolved":
        tickets = tickets.filter(status__in=["resolved", "closed"])

    if search_q:
        tickets = tickets.filter(
            Q(subject__icontains=search_q) | Q(description__icontains=search_q) | Q(category__icontains=search_q)
        )

    form = SupportTicketForm()
    return render(
        request,
        "attendance/my_tickets.html",
        {
            "tickets": tickets,
            "open_count": all_tickets.filter(status__in=["open", "in_progress", "waiting"]).count(),
            "resolved_count": all_tickets.filter(status__in=["resolved", "closed"]).count(),
            "filter_status": filter_status,
            "search_q": search_q,
            "form": form,
            "employee": employee,
        },
    )


# ─── ATTENDANCE CORRECTION REQUESTS (Employee) ───────────────────────────────


@login_required
def request_correction(request):
    try:
        employee = request.user.employee
    except Employee.DoesNotExist:
        return redirect("dashboard")

    if request.method == "POST":
        form = AttendanceCorrectionForm(request.POST)
        if form.is_valid():
            correction = form.save(commit=False)
            correction.employee = employee
            correction.save()
            email_service.notify_admin_correction_raised(correction)
            messages.success(request, "Correction request submitted! Admin will review it soon.")
            return redirect("my_corrections")
    else:
        form = AttendanceCorrectionForm(initial={"date": timezone.now().date()})

    recent = AttendanceCorrectionRequest.objects.filter(employee=employee)[:5]
    return render(
        request,
        "attendance/request_correction.html",
        {
            "form": form,
            "employee": employee,
            "recent": recent,
        },
    )


@login_required
def my_corrections(request):
    try:
        employee = request.user.employee
    except Employee.DoesNotExist:
        return redirect("dashboard")

    corrections = AttendanceCorrectionRequest.objects.filter(employee=employee)
    return render(
        request,
        "attendance/my_corrections.html",
        {
            "corrections": corrections,
            "pending_count": corrections.filter(status="pending").count(),
            "approved_count": corrections.filter(status="approved").count(),
            "rejected_count": corrections.filter(status="rejected").count(),
            "employee": employee,
        },
    )


# ─── ADMIN LEAVE & TICKET VIEWS ──────────────────────────────────────────────


@login_required
@user_passes_test(is_admin)
def admin_leaves(request):
    status_filter = request.GET.get("status", "")
    leaves = LeaveRequest.objects.select_related("employee__user", "employee__department").all()
    if status_filter:
        leaves = leaves.filter(status=status_filter)

    return render(
        request,
        "attendance/admin_leaves.html",
        {
            "leaves": leaves,
            "status_filter": status_filter,
            "pending_count": LeaveRequest.objects.filter(status="pending").count(),
        },
    )


@login_required
@user_passes_test(is_admin)
def admin_leave_action(request, leave_id):
    leave = get_object_or_404(LeaveRequest, id=leave_id)
    if request.method == "POST":
        form = AdminLeaveResponseForm(request.POST, instance=leave)
        if form.is_valid():
            prev_status = leave.status
            updated_leave = form.save()

            if updated_leave.status != prev_status:
                if updated_leave.status in ("approved", "rejected"):
                    email_service.notify_employee_leave_decision(updated_leave)

                if updated_leave.status == "approved":
                    # Create leave attendance records for every Mon–Sat in the range
                    cur = updated_leave.from_date
                    while cur <= updated_leave.to_date:
                        if cur.weekday() != 6:  # skip Sunday
                            AttendanceRecord.objects.get_or_create(
                                employee=updated_leave.employee,
                                date=cur,
                                defaults={"status": "leave"},
                            )
                        cur += datetime.timedelta(days=1)

                elif prev_status == "approved":
                    # Leave was revoked — remove auto-created leave records (only those with no check-in)
                    cur = updated_leave.from_date
                    while cur <= updated_leave.to_date:
                        if cur.weekday() != 6:  # skip Sunday (was never created, but be consistent)
                            AttendanceRecord.objects.filter(
                                employee=updated_leave.employee,
                                date=cur,
                                status="leave",
                                check_in_time__isnull=True,
                            ).delete()
                        cur += datetime.timedelta(days=1)

            messages.success(request, f"Leave request {updated_leave.get_status_display()} successfully.")
            return redirect("admin_leaves")
    else:
        form = AdminLeaveResponseForm(instance=leave)

    return render(request, "attendance/admin_leave_action.html", {"form": form, "leave": leave})


@login_required
@user_passes_test(is_admin)
def admin_tickets(request):
    status_filter = request.GET.get("status", "")
    tickets = SupportTicket.objects.select_related("employee__user", "employee__department").all()
    if status_filter:
        tickets = tickets.filter(status=status_filter)

    if request.method == "POST":
        ids = request.POST.getlist("ticket_ids")
        if ids:
            deleted, _ = SupportTicket.objects.filter(id__in=ids).delete()
            messages.success(request, f"{deleted} ticket(s) deleted.")
        return redirect(f"{request.path}{'?status=' + status_filter if status_filter else ''}")

    return render(
        request,
        "attendance/admin_tickets.html",
        {
            "tickets": tickets,
            "status_filter": status_filter,
            "open_count": SupportTicket.objects.filter(status="open").count(),
        },
    )


@login_required
@user_passes_test(is_admin)
def admin_ticket_action(request, ticket_id):
    ticket = get_object_or_404(SupportTicket, id=ticket_id)
    if request.method == "POST":
        form = AdminTicketResponseForm(request.POST, instance=ticket)
        if form.is_valid():
            prev_status = ticket.status
            updated_ticket = form.save()
            if updated_ticket.status != prev_status or updated_ticket.admin_response:
                email_service.notify_employee_ticket_updated(updated_ticket)
            messages.success(request, f"Ticket #{updated_ticket.id} updated successfully.")
            return redirect("admin_tickets")
    else:
        form = AdminTicketResponseForm(instance=ticket)

    return render(request, "attendance/admin_ticket_action.html", {"form": form, "ticket": ticket})


# ─── ADMIN CORRECTION VIEWS ──────────────────────────────────────────────────


@login_required
@user_passes_test(is_admin)
def admin_corrections(request):
    status_filter = request.GET.get("status", "pending")
    corrections = AttendanceCorrectionRequest.objects.select_related("employee__user").all()
    if status_filter:
        corrections = corrections.filter(status=status_filter)

    return render(
        request,
        "attendance/admin_corrections.html",
        {
            "corrections": corrections,
            "status_filter": status_filter,
            "pending_count": AttendanceCorrectionRequest.objects.filter(status="pending").count(),
        },
    )


@login_required
@user_passes_test(is_admin)
def admin_correction_action(request, correction_id):
    correction = get_object_or_404(AttendanceCorrectionRequest, id=correction_id)

    if request.method == "POST":
        form = AdminCorrectionResponseForm(request.POST, instance=correction)
        if form.is_valid():
            saved = form.save()

            if saved.status == "approved":
                record, _ = AttendanceRecord.objects.get_or_create(
                    employee=saved.employee, date=saved.date, defaults={"status": "present"}
                )
                if saved.requested_check_in:
                    record.check_in_time = saved.requested_check_in
                if saved.requested_check_out:
                    record.check_out_time = saved.requested_check_out
                record.save()
                if record.check_in_time and record.check_out_time:
                    # Both times present — calculate present / half_day based on hours
                    record.calculate_hours()
                elif record.check_in_time:
                    # Only check-in corrected (admin reviewed and approved) → present
                    record.status = "present"
                    record.save()
                messages.success(
                    request,
                    f"Approved — attendance for {saved.employee.user.get_full_name()} "
                    f"on {saved.date} has been updated.",
                )
            else:
                messages.info(request, f"Correction request for {saved.employee.user.get_full_name()} rejected.")
            email_service.notify_employee_correction_decision(saved)
            return redirect("admin_corrections")
    else:
        form = AdminCorrectionResponseForm(instance=correction)

    return render(
        request,
        "attendance/admin_correction_action.html",
        {
            "form": form,
            "correction": correction,
        },
    )


# ─── SALARY VIEWS ────────────────────────────────────────────────────────────


@login_required
@user_passes_test(is_admin)
def admin_salary(request):
    """Admin sees all employees' salary for a selected month/year."""
    import calendar

    _resolve_incomplete_checkins()
    month = int(request.GET.get("month", timezone.now().month))
    year = int(request.GET.get("year", timezone.now().year))
    status_filter = request.GET.get("filter", "total")

    employees = Employee.objects.filter(is_active=True).select_related("user", "department")
    salary_map = {s.employee_id: s for s in SalaryRecord.objects.filter(month=month, year=year)}

    all_rows = []
    has_unprocessed = False
    for emp in employees:
        sal = salary_map.get(emp.id)
        structure = None
        if not sal:
            structure = EmployeeSalaryStructure.active_for(emp, month, year)
            if not structure:
                # Fallback: employee has past SalaryRecords but no structure yet
                # (set before auto-structure feature). Use last non-zero record as preview.
                last_record = (
                    SalaryRecord.objects.filter(employee=emp, basic_salary__gt=0).order_by("-year", "-month").first()
                )
                if last_record:
                    import datetime as _dt

                    effective = _dt.date(last_record.year, last_record.month, 1)
                    structure, _ = EmployeeSalaryStructure.objects.get_or_create(
                        employee=emp,
                        effective_from=effective,
                        defaults={
                            "basic_salary": last_record.basic_salary,
                            "allowances": last_record.allowances,
                            "notes": f"Auto-migrated from {calendar.month_name[last_record.month]} {last_record.year}",
                        },
                    )
            if structure:
                has_unprocessed = True
        all_rows.append({"employee": emp, "salary": sal, "structure": structure})

    paid_count = sum(1 for r in all_rows if r["salary"] and r["salary"].is_paid)
    unpaid_count = len(all_rows) - paid_count

    if status_filter == "paid":
        rows = [r for r in all_rows if r["salary"] and r["salary"].is_paid]
    elif status_filter == "pending":
        rows = [r for r in all_rows if not r["salary"] or not r["salary"].is_paid]
    else:
        rows = all_rows

    return render(
        request,
        "attendance/admin_salary.html",
        {
            "rows": rows,
            "month": month,
            "year": year,
            "month_name": calendar.month_name[month],
            "paid_count": paid_count,
            "unpaid_count": unpaid_count,
            "total": len(all_rows),
            "status_filter": status_filter,
            "has_unprocessed": has_unprocessed,
            "months": [(i, calendar.month_name[i]) for i in range(1, 13)],
            "years": range(2023, timezone.now().year + 1),
        },
    )


@login_required
@user_passes_test(is_admin)
def process_all_salaries(request):
    """One-click: auto-calculate and create salary records for all employees from their structure."""
    import calendar

    if request.method != "POST":
        return redirect("admin_salary")

    _resolve_incomplete_checkins()
    month = int(request.POST.get("month", timezone.now().month))
    year = int(request.POST.get("year", timezone.now().year))
    last_day_of_month = calendar.monthrange(year, month)[1]
    today_date = timezone.now().date()
    cutoff_day = today_date.day if (year == today_date.year and month == today_date.month) else last_day_of_month

    employees = Employee.objects.filter(is_active=True)
    processed = 0
    skipped_no_structure = 0
    skipped_paid = 0

    for employee in employees:
        active_structure = EmployeeSalaryStructure.active_for(employee, month, year)
        if not active_structure:
            last_record = (
                SalaryRecord.objects.filter(employee=employee, basic_salary__gt=0).order_by("-year", "-month").first()
            )
            if last_record:
                effective = datetime.date(last_record.year, last_record.month, 1)
                active_structure, _ = EmployeeSalaryStructure.objects.get_or_create(
                    employee=employee,
                    effective_from=effective,
                    defaults={
                        "basic_salary": last_record.basic_salary,
                        "allowances": last_record.allowances,
                        "notes": f"Auto-migrated from {calendar.month_name[last_record.month]} {last_record.year}",
                    },
                )
        if not active_structure:
            skipped_no_structure += 1
            continue

        # Skip already-paid records
        existing = SalaryRecord.objects.filter(employee=employee, month=month, year=year).first()
        if existing and existing.is_paid:
            skipped_paid += 1
            continue

        # Attendance calculation (same formula as update_salary)
        month_records = AttendanceRecord.objects.filter(employee=employee, date__month=month, date__year=year)
        present_days = sum(1 for r in month_records if r.status == "present" and r.date.weekday() != 6)
        half_days = sum(1 for r in month_records if r.status == "half_day" and r.date.weekday() != 6)
        leave_days = sum(1 for r in month_records if r.status == "leave" and r.date.weekday() != 6)
        total_working_days = sum(1 for d in range(1, cutoff_day + 1) if datetime.date(year, month, d).weekday() != 6)
        sunday_days = sum(1 for d in range(1, cutoff_day + 1) if datetime.date(year, month, d).weekday() == 6)
        absent_days = max(0, total_working_days - present_days - half_days - leave_days)

        # Salary formula
        basic_val = float(active_structure.basic_salary)
        _rate = basic_val / last_day_of_month if last_day_of_month else 0
        present_earn = round(present_days * _rate, 2)
        sunday_earn = round(sunday_days * _rate, 2)
        half_earn = round(half_days * _rate * 0.5, 2)
        leave_deduct = round(leave_days * _rate, 2)
        absent_deduct = round(absent_days * _rate, 2)
        half_deduct = round(half_days * _rate * 0.5, 2)
        total_deductions = round(leave_deduct + absent_deduct + half_deduct, 2)
        computed_net = round(present_earn + sunday_earn + half_earn + float(active_structure.allowances), 2)

        salary, _ = SalaryRecord.objects.get_or_create(
            employee=employee,
            month=month,
            year=year,
            defaults={
                "basic_salary": active_structure.basic_salary,
                "allowances": active_structure.allowances,
                "absent_days": absent_days,
                "half_days": half_days,
            },
        )
        salary.basic_salary = active_structure.basic_salary
        salary.allowances = active_structure.allowances
        salary.absent_days = absent_days
        salary.half_days = half_days
        salary._skip_auto_calc = True
        salary.deductions = total_deductions
        salary.net_salary = computed_net
        salary.save()
        processed += 1

    parts = [f"Processed {processed} employees."]
    if skipped_paid:
        parts.append(f"{skipped_paid} already paid (skipped).")
    if skipped_no_structure:
        parts.append(f"{skipped_no_structure} have no salary structure set.")
    messages.success(request, " ".join(parts))
    return redirect(f"/admin-panel/salary/?month={month}&year={year}")


@login_required
@user_passes_test(is_admin)
def update_salary(request, emp_id):
    import calendar

    employee = get_object_or_404(Employee, id=emp_id)
    month = int(request.GET.get("month", timezone.now().month))
    year = int(request.GET.get("year", timezone.now().year))

    # ── Active salary structure for this month ────────────────────────────
    active_structure = EmployeeSalaryStructure.active_for(employee, month, year)

    # ── Get or create payroll record, auto-fill from structure ────────────
    with transaction.atomic():
        salary, created = SalaryRecord.objects.get_or_create(
            employee=employee,
            month=month,
            year=year,
            defaults={
                "basic_salary": active_structure.basic_salary if active_structure else 0,
                "allowances": active_structure.allowances if active_structure else 0,
                "absent_days": 0,
            },
        )

    # ── Attendance counts ─────────────────────────────────────────────────
    month_records = AttendanceRecord.objects.filter(employee=employee, date__month=month, date__year=year)
    present_days = sum(1 for r in month_records if r.status == "present" and r.date.weekday() != 6)
    half_days = sum(1 for r in month_records if r.status == "half_day" and r.date.weekday() != 6)
    leave_days = sum(1 for r in month_records if r.status == "leave" and r.date.weekday() != 6)

    # Cutoff: current month → elapsed days only; past/future → full month
    today_date = timezone.now().date()
    last_day_of_month = calendar.monthrange(year, month)[1]
    cutoff_day = today_date.day if (year == today_date.year and month == today_date.month) else last_day_of_month

    # Mon–Sat working days in period
    total_working_days = sum(1 for d in range(1, cutoff_day + 1) if datetime.date(year, month, d).weekday() != 6)
    # Sundays in period — paid holiday, always included in salary
    sunday_days = sum(1 for d in range(1, cutoff_day + 1) if datetime.date(year, month, d).weekday() == 6)

    # Absent = Mon–Sat days with no record and not on approved leave
    absent_days = max(0, total_working_days - present_days - half_days - leave_days)

    # ── Salary formula ────────────────────────────────────────────────────
    # per_day = basic / actual_days_in_month  (NOT 30.4)
    # This ensures: payable + deductions = basic (exact, no rounding gap)
    #
    # Payable     = (present + sunday) × rate + half × rate × 0.5
    # Deductions  = (leave + absent + half × 0.5) × rate  [shown only]
    # Net salary  = Payable + allowances
    basic_val = float(salary.basic_salary)
    _rate = basic_val / last_day_of_month if last_day_of_month else 0
    daily_rate = round(_rate, 2)  # display only
    present_earn = round(present_days * _rate, 2)
    sunday_earn = round(sunday_days * _rate, 2)
    half_earn = round(half_days * _rate * 0.5, 2)
    leave_deduct = round(leave_days * _rate, 2)
    absent_deduct = round(absent_days * _rate, 2)
    half_deduct = round(half_days * _rate * 0.5, 2)
    total_deductions = round(leave_deduct + absent_deduct + half_deduct, 2)
    # Net = payable only (deductions are informational, not subtracted)
    computed_net = round(present_earn + sunday_earn + half_earn + float(salary.allowances), 2)

    salary.absent_days = absent_days
    salary.half_days = half_days
    salary._skip_auto_calc = True
    salary.deductions = total_deductions
    salary.net_salary = computed_net
    salary.save()

    if request.method == "POST":
        form = SalaryForm(request.POST, instance=salary)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.absent_days = absent_days
            obj.half_days = half_days
            # Recalculate with updated basic/allowances (same formula as GET)
            r = float(obj.basic_salary) / last_day_of_month if last_day_of_month else 0
            p_earn = round(present_days * r, 2)
            s_earn = round(sunday_days * r, 2)
            h_earn = round(half_days * r * 0.5, 2)
            l_ded = round(leave_days * r, 2)
            a_ded = round(absent_days * r, 2)
            h_ded = round(half_days * r * 0.5, 2)
            t_ded = round(l_ded + a_ded + h_ded, 2)
            new_net = round(p_earn + s_earn + h_earn + float(obj.allowances), 2)
            obj._skip_auto_calc = True
            obj.deductions = t_ded
            obj.net_salary = new_net
            obj.save()

            # ── Auto-save salary structure so future months use this salary ──
            # Only create a new structure entry if salary changed or no structure exists
            new_basic = obj.basic_salary
            new_allowances = obj.allowances
            effective_date = datetime.date(year, month, 1)
            structure_changed = (
                active_structure is None
                or active_structure.basic_salary != new_basic
                or active_structure.allowances != new_allowances
            )
            if structure_changed:
                EmployeeSalaryStructure.objects.update_or_create(
                    employee=employee,
                    effective_from=effective_date,
                    defaults={
                        "basic_salary": new_basic,
                        "allowances": new_allowances,
                        "notes": f"Set from payroll — {calendar.month_name[month]} {year}",
                    },
                )

            messages.success(
                request,
                f"Salary for {employee.user.get_full_name()} "
                f"({calendar.month_name[month]} {year}) saved and set as default for future months.",
            )
            return redirect(f"{request.path_info}?month={month}&year={year}&saved=1")
    else:
        form = SalaryForm(instance=salary)

    return render(
        request,
        "attendance/update_salary.html",
        {
            "form": form,
            "employee": employee,
            "salary": salary,
            "month": month,
            "year": year,
            "month_name": calendar.month_name[month],
            "present_days": present_days,
            "sunday_days": sunday_days,
            "half_days": half_days,
            "leave_days": leave_days,
            "absent_days": absent_days,
            "present_earn": present_earn,
            "sunday_earn": sunday_earn,
            "half_earn": half_earn,
            "leave_deduct": leave_deduct,
            "absent_deduct": absent_deduct,
            "half_deduct": half_deduct,
            "total_deductions": total_deductions,
            "daily_rate": daily_rate,
            "days_in_month": last_day_of_month,
            "active_structure": active_structure,
            "auto_filled": created and active_structure is not None,
            "months": [(i, calendar.month_name[i]) for i in range(1, 13)],
            "years": range(2023, timezone.now().year + 1),
            "saved": request.GET.get("saved"),
        },
    )


@login_required
@user_passes_test(is_admin)
def salary_structure(request, emp_id):
    """View/update an employee's salary structure with full history."""
    import calendar as _cal

    employee = get_object_or_404(Employee, id=emp_id)
    today = timezone.now().date()

    if request.method == "POST":
        form = SalaryStructureForm(request.POST)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.employee = employee
            EmployeeSalaryStructure.objects.update_or_create(
                employee=employee,
                effective_from=obj.effective_from,
                defaults={"basic_salary": obj.basic_salary, "allowances": obj.allowances, "notes": obj.notes},
            )
            messages.success(
                request,
                f"Salary updated for {employee.user.get_full_name()} — "
                f"₹{obj.basic_salary:,.0f} effective {_cal.month_name[obj.effective_from.month]} {obj.effective_from.year}.",
            )
            return redirect("update_salary", emp_id=emp_id)
    else:
        current = EmployeeSalaryStructure.active_for(employee, today.month, today.year)
        form = SalaryStructureForm(
            initial={
                "basic_salary": current.basic_salary if current else 0,
                "allowances": current.allowances if current else 0,
                "effective_from": today.replace(day=1),
            }
        )

    history = EmployeeSalaryStructure.objects.filter(employee=employee)
    return render(
        request,
        "attendance/salary_structure.html",
        {
            "form": form,
            "employee": employee,
            "history": history,
        },
    )


@login_required
def my_salary(request):
    """Employee sees their own salary slips."""
    import calendar

    try:
        employee = request.user.employee
    except Employee.DoesNotExist:
        return redirect("dashboard")

    salaries = SalaryRecord.objects.filter(employee=employee)

    # Current month salary highlighted
    current_month = timezone.now().month
    current_year = timezone.now().year
    current_salary = salaries.filter(month=current_month, year=current_year).first()

    return render(
        request,
        "attendance/my_salary.html",
        {
            "employee": employee,
            "salaries": salaries,
            "current_salary": current_salary,
            "current_month": calendar.month_name[current_month],
            "current_year": current_year,
        },
    )


@login_required
def salary_slip(request, salary_id):
    import calendar

    slip = get_object_or_404(SalaryRecord, id=salary_id)
    # Access control: admin can view any, employee only their own
    if not request.user.is_staff:
        try:
            if slip.employee != request.user.employee:
                messages.error(request, "Access denied.")
                return redirect("my_salary")
        except Employee.DoesNotExist:
            return redirect("my_salary")
    daily_rate = round(slip.daily_rate, 2)
    total_earnings = round(float(slip.basic_salary) + float(slip.allowances), 2)
    absent_deduction = round(slip.absent_days * slip.daily_rate, 2)
    half_day_deduction = round(slip.half_days * (slip.daily_rate / 2), 2)
    return render(
        request,
        "attendance/salary_slip.html",
        {
            "slip": slip,
            "employee": slip.employee,
            "month_name": calendar.month_name[slip.month],
            "daily_rate": daily_rate,
            "total_earnings": total_earnings,
            "absent_deduction": absent_deduction,
            "half_day_deduction": half_day_deduction,
        },
    )


# ─── EMAIL ONE-CLICK ACTION ──────────────────────────────────────────────────


# ─── REIMBURSEMENT VIEWS ────────────────────────────────────────────────────


@login_required
def my_reimbursements(request):
    if request.user.is_staff and get_panel_mode(request) == "admin":
        return redirect("admin_reimbursements")
    employee = get_object_or_404(Employee, user=request.user)
    status_filter = request.GET.get("status", "")
    qs = employee.reimbursements.prefetch_related("attachments").all()
    if status_filter in ("pending", "approved", "rejected"):
        qs = qs.filter(status=status_filter)

    total_approved = employee.reimbursements.filter(status="approved").aggregate(total=Sum("amount"))["total"] or 0
    pending_count = employee.reimbursements.filter(status="pending").count()
    approved_count = employee.reimbursements.filter(status="approved").count()
    rejected_count = employee.reimbursements.filter(status="rejected").count()

    return render(
        request,
        "attendance/my_reimbursements.html",
        {
            "reimbursements": qs,
            "status_filter": status_filter,
            "total_approved": total_approved,
            "pending_count": pending_count,
            "approved_count": approved_count,
            "rejected_count": rejected_count,
            "today": timezone.now().date(),
        },
    )


@login_required
def create_reimbursement(request):
    if request.user.is_staff and get_panel_mode(request) == "admin":
        return redirect("admin_reimbursements")
    employee = get_object_or_404(Employee, user=request.user)

    if request.method == "POST":
        form = ReimbursementForm(request.POST, request.FILES)
        files = request.FILES.getlist("attachments")

        if form.is_valid():
            if not files:
                messages.error(
                    request, "Please upload at least one supporting document (invoice, receipt, or screenshot)."
                )
            else:
                allowed_exts = {".pdf", ".jpg", ".jpeg", ".png", ".gif"}
                max_size = 5 * 1024 * 1024  # 5 MB
                file_errors = []
                for f in files:
                    ext = os.path.splitext(f.name)[1].lower()
                    if ext not in allowed_exts:
                        file_errors.append(f"'{f.name}': unsupported type. Allowed: PDF, JPG, PNG, GIF.")
                    elif f.size > max_size:
                        file_errors.append(f"'{f.name}' exceeds the 5 MB size limit.")

                if file_errors:
                    for err in file_errors:
                        messages.error(request, err)
                else:
                    reimbursement = form.save(commit=False)
                    reimbursement.employee = employee
                    reimbursement.save()

                    for f in files:
                        ReimbursementAttachment.objects.create(
                            reimbursement=reimbursement,
                            file=f,
                            filename=f.name,
                        )

                    email_service.notify_admin_reimbursement_submitted(reimbursement)
                    email_service.notify_employee_reimbursement_submitted(reimbursement)

                    messages.success(
                        request,
                        f"Reimbursement request #{reimbursement.id} submitted successfully. You'll be notified once reviewed.",
                    )
                    return redirect("my_reimbursements")
    else:
        form = ReimbursementForm(initial={"expense_date": timezone.now().date()})

    return render(
        request,
        "attendance/create_reimbursement.html",
        {"form": form, "today": timezone.now().date()},
    )


@login_required
@user_passes_test(is_admin)
def bulk_reimb_action(request):
    if request.method != "POST":
        return redirect("admin_reimbursements")
    ids = request.POST.getlist("reimb_ids")
    action = request.POST.get("action", "")
    next_url = request.POST.get("next", "")
    if ids and action in ("approved", "rejected"):
        updated = Reimbursement.objects.filter(id__in=ids).update(
            status=action,
            reviewed_by=request.user,
            reviewed_at=timezone.now(),
        )
        label = "approved" if action == "approved" else "rejected"
        messages.success(request, f"{updated} reimbursement(s) {label}.")
    if next_url and next_url.startswith("/"):
        return redirect(next_url)
    return redirect("admin_reimbursements")


@login_required
@user_passes_test(is_admin)
def admin_reimbursements(request):
    qs = Reimbursement.objects.select_related("employee__user", "employee__department", "reviewed_by").prefetch_related(
        "attachments"
    )

    status_filter = request.GET.get("status", "")
    employee_filter = request.GET.get("employee", "").strip()
    category_filter = request.GET.get("category", "")

    if status_filter in ("pending", "approved", "rejected"):
        qs = qs.filter(status=status_filter)
    if employee_filter:
        try:
            qs = qs.filter(employee_id=int(employee_filter))
        except ValueError:
            pass
    if category_filter:
        qs = qs.filter(category=category_filter)

    all_reimb = Reimbursement.objects.all()
    total_count = all_reimb.count()
    pending_count = all_reimb.filter(status="pending").count()
    approved_count = all_reimb.filter(status="approved").count()
    rejected_count = all_reimb.filter(status="rejected").count()

    employees = Employee.objects.filter(is_active=True).select_related("user").order_by("user__first_name")

    return render(
        request,
        "attendance/admin_reimbursements.html",
        {
            "reimbursements": qs,
            "total_count": total_count,
            "pending_count": pending_count,
            "approved_count": approved_count,
            "rejected_count": rejected_count,
            "status_filter": status_filter,
            "employee_filter": employee_filter,
            "category_filter": category_filter,
            "category_choices": Reimbursement.CATEGORY_CHOICES,
            "employees": employees,
            "today": timezone.now().date(),
        },
    )


@login_required
@user_passes_test(is_admin)
def admin_reimbursement_detail(request, reimbursement_id):
    reimbursement = get_object_or_404(
        Reimbursement.objects.select_related("employee__user", "employee__department", "reviewed_by").prefetch_related(
            "attachments"
        ),
        id=reimbursement_id,
    )

    if request.method == "POST":
        form = AdminReimbursementActionForm(request.POST, instance=reimbursement)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.reviewed_by = request.user
            obj.reviewed_at = timezone.now()
            obj.save()
            email_service.notify_employee_reimbursement_decision(reimbursement)
            action_label = "approved" if reimbursement.status == "approved" else "rejected"
            messages.success(request, f"Reimbursement #{reimbursement.id} has been {action_label}. Employee notified.")
            return redirect("admin_reimbursements")
    else:
        form = AdminReimbursementActionForm(instance=reimbursement)

    return render(
        request,
        "attendance/admin_reimbursement_detail.html",
        {
            "reimbursement": reimbursement,
            "form": form,
            "today": timezone.now().date(),
        },
    )


@login_required
@user_passes_test(is_admin)
def export_reimbursements_excel(request):
    import csv

    qs = Reimbursement.objects.select_related("employee__user", "employee__department", "reviewed_by").order_by(
        "-submitted_at"
    )

    status_filter = request.GET.get("status", "")
    if status_filter in ("pending", "approved", "rejected"):
        qs = qs.filter(status=status_filter)

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="reimbursements.csv"'
    writer = csv.writer(response)
    writer.writerow(
        [
            "ID",
            "Employee",
            "Employee ID",
            "Department",
            "Title",
            "Category",
            "Amount",
            "Expense Date",
            "Payment Method",
            "Status",
            "Submitted On",
            "Reviewed By",
            "Reviewed On",
            "Admin Remarks",
        ]
    )
    for r in qs:
        writer.writerow(
            [
                r.id,
                r.employee.user.get_full_name(),
                r.employee.employee_id,
                str(r.employee.department) if r.employee.department else "",
                r.title,
                r.get_category_display(),
                r.amount,
                r.expense_date.strftime("%d %b %Y"),
                r.get_payment_method_display(),
                r.get_status_display(),
                r.submitted_at.strftime("%d %b %Y %H:%M"),
                r.reviewed_by.get_full_name() if r.reviewed_by else "",
                r.reviewed_at.strftime("%d %b %Y %H:%M") if r.reviewed_at else "",
                r.admin_remarks,
            ]
        )
    return response


def email_action(request, token):
    """
    Handle approve/reject/resolve actions triggered from email buttons.
    No login required — the signed token is the authentication.
    Token expires after 7 days.
    """
    from django.core import signing

    try:
        data = email_service.decode_action_token(token)
    except signing.SignatureExpired:
        return render(
            request,
            "attendance/email_action_result.html",
            {
                "success": False,
                "title": "Link Expired",
                "message": "This action link has expired (valid for 7 days). Please log in to the admin panel to take action.",
            },
        )
    except signing.BadSignature:
        return render(
            request,
            "attendance/email_action_result.html",
            {
                "success": False,
                "title": "Invalid Link",
                "message": "This link is invalid or has been tampered with.",
            },
        )

    obj_type = data["t"]
    obj_id = data["id"]
    action = data["a"]

    # ── Leave ────────────────────────────────────────────────────────────────
    if obj_type == "leave":
        leave = get_object_or_404(LeaveRequest, id=obj_id)

        if leave.status != "pending":
            return render(
                request,
                "attendance/email_action_result.html",
                {
                    "success": False,
                    "title": "Already Actioned",
                    "message": f"This leave request was already {leave.status}. No changes made.",
                    "employee_name": leave.employee.user.get_full_name(),
                },
            )

        leave.status = action  # 'approved' or 'rejected'
        leave.save()

        if action == "approved":
            cur = leave.from_date
            while cur <= leave.to_date:
                if cur.weekday() != 6:
                    AttendanceRecord.objects.get_or_create(
                        employee=leave.employee,
                        date=cur,
                        defaults={"status": "leave"},
                    )
                cur += datetime.timedelta(days=1)

        email_service.notify_employee_leave_decision(leave)

        action_label = "Approved" if action == "approved" else "Rejected"
        return render(
            request,
            "attendance/email_action_result.html",
            {
                "success": True,
                "title": f"Leave {action_label}",
                "action_label": action_label,
                "obj_type": "Leave Request",
                "employee_name": leave.employee.user.get_full_name(),
                "detail": f'{leave.from_date.strftime("%d %b %Y")} → {leave.to_date.strftime("%d %b %Y")} · {leave.total_days} working day(s)',
                "message": f"Leave has been {action_label.lower()}. The employee has been notified by email.",
            },
        )

    # ── Ticket ───────────────────────────────────────────────────────────────
    elif obj_type == "ticket":
        ticket = get_object_or_404(SupportTicket, id=obj_id)

        if ticket.status in ("resolved", "closed"):
            return render(
                request,
                "attendance/email_action_result.html",
                {
                    "success": False,
                    "title": "Already Actioned",
                    "message": f"Ticket #{ticket.id} is already {ticket.status}. No changes made.",
                    "employee_name": ticket.employee.user.get_full_name(),
                },
            )

        ticket.status = action  # 'resolved' or 'in_progress'
        ticket.save()
        email_service.notify_employee_ticket_updated(ticket)

        action_label = "Resolved" if action == "resolved" else "Marked In Progress"
        return render(
            request,
            "attendance/email_action_result.html",
            {
                "success": True,
                "title": f"Ticket {action_label}",
                "action_label": action_label,
                "obj_type": f"Ticket #{ticket.id}",
                "employee_name": ticket.employee.user.get_full_name(),
                "detail": ticket.subject,
                "message": f"Ticket #{ticket.id} has been {action_label.lower()}. The employee has been notified by email.",
            },
        )

    # ── Correction ───────────────────────────────────────────────────────────
    elif obj_type == "correction":
        correction = get_object_or_404(AttendanceCorrectionRequest, id=obj_id)

        if correction.status != "pending":
            return render(
                request,
                "attendance/email_action_result.html",
                {
                    "success": False,
                    "title": "Already Actioned",
                    "message": f"This correction request was already {correction.status}. No changes made.",
                    "employee_name": correction.employee.user.get_full_name(),
                },
            )

        correction.status = action  # 'approved' or 'rejected'
        correction.save()

        if action == "approved":
            record, _ = AttendanceRecord.objects.get_or_create(
                employee=correction.employee,
                date=correction.date,
                defaults={"status": "present"},
            )
            if correction.requested_check_in:
                record.check_in_time = correction.requested_check_in
            if correction.requested_check_out:
                record.check_out_time = correction.requested_check_out
            record.save()
            record.calculate_hours()

        email_service.notify_employee_correction_decision(correction)

        action_label = "Approved" if action == "approved" else "Rejected"
        return render(
            request,
            "attendance/email_action_result.html",
            {
                "success": True,
                "title": f"Correction {action_label}",
                "action_label": action_label,
                "obj_type": "Correction Request",
                "employee_name": correction.employee.user.get_full_name(),
                "detail": f'Date: {correction.date.strftime("%d %b %Y")}',
                "message": f"Correction request has been {action_label.lower()}. The employee has been notified by email.",
            },
        )

    # ── Reimbursement ────────────────────────────────────────────────────────
    elif obj_type == "reimbursement":
        reimbursement = get_object_or_404(Reimbursement, id=obj_id)

        if reimbursement.status != "pending":
            return render(
                request,
                "attendance/email_action_result.html",
                {
                    "success": False,
                    "title": "Already Actioned",
                    "message": f"Reimbursement #{reimbursement.id} was already {reimbursement.status}. No changes made.",
                    "employee_name": reimbursement.employee.user.get_full_name(),
                },
            )

        reimbursement.status = action  # 'approved' or 'rejected'
        reimbursement.reviewed_at = timezone.now()
        reimbursement.save()
        email_service.notify_employee_reimbursement_decision(reimbursement)

        action_label = "Approved" if action == "approved" else "Rejected"
        return render(
            request,
            "attendance/email_action_result.html",
            {
                "success": True,
                "title": f"Reimbursement {action_label}",
                "action_label": action_label,
                "obj_type": f"Reimbursement #{reimbursement.id}",
                "employee_name": reimbursement.employee.user.get_full_name(),
                "detail": f"{reimbursement.title} · ₹{reimbursement.amount}",
                "message": f"Reimbursement #{reimbursement.id} has been {action_label.lower()}. The employee has been notified by email.",
            },
        )

    return render(
        request,
        "attendance/email_action_result.html",
        {
            "success": False,
            "title": "Unknown Action",
            "message": "Unrecognised action type in this link.",
        },
    )
